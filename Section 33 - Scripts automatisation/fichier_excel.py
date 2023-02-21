# 234
# Fichiers Excels: Premiers pas

import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx")

sheet_name = wb.sheetnames # permet de récupérer le nom des feuilles

sheet = wb[wb.sheetnames[0]] # permet de récupérer le nom de la première feuille

sheet2 = wb.active # permet de récupérer la page active

cell = sheet["A1"].value # permet de récupérer la valeur de la case A1

cell2 = sheet2.cell(4, 3)

""" 
# permet de récup toutes les valeurs d'une colonne
for row in range(2, 7):
    cell3 = sheet2.cell(row, 2)
    print(cell3.value) 
"""

print(sheet2.max_column) # permet de récup le nb de colonnes
print(sheet2.max_row) # permet de récup le nb de lignes

